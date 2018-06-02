#! /usr/bin/env python3
# -*- coding:utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

#读取文件和相应的表
wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    combine_sheet = wb.create_sheet(title='combine')#创建表和相应的表头
    combine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])
    for stu in students_sheet.values:    #合并表格并以此添加到combine表中
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    combine_sheet.append(list(stu) + [time[2]])
    wb.save('combine.xlsx')

def split():
    combine_sheet = wb['combine']
    split_name = [] #用来存储表中的年份
    for item in combine_sheet.values:
        if item[0] != '创建时间':
            split_name.append(item[0].strftime("%Y"))
    
    #分别存储数据
    for name in set(split_name):
        #创建文件
        wb_temp = Workbook()
        #删除已有的表
        wb_temp.remove(wb_temp.active)
        ws = wb_temp.create_sheet(title=name)
        #写入相应年份的数据
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0] == name:
                    ws.append(item_by_year)
        #覆盖保存
        wb_temp.save('{}.xlsx'.format(name))

if __name__ == '__main__':
    combine()
    split()
